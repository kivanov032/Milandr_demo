import cv2
import numpy as np
import json
import datetime
import shutil
import time
from pathlib import Path
import atexit
from typing import Optional, Tuple, Dict
from threading import Lock
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from project.application.addition.exceptions import ProtocolException
from project.application.addition.logger import logger
from project.application.data_work.wafer_data import WaferMap, DieStatus
from project.application.data_work.wafer_map_bin_patcher import WaferMapBinPatcher
from project.configuration.worker import read_from_json


class Protocol:
    """
    Класс для хранения названий папок и файлов протоколов и их генерации (Singleton).

    Attributes:
        _instance (Optional['Protocol']): Единственный экземпляр класса
        _lock (Lock): Блокировка для потокобезопасности
        _file_lock (Lock): Блокировка для файловых операций
        _main_folder_path (Optional[Path]): Путь к папке конкретного кристалла
        _before_AOI_folder_path (Optional[Path]): Путь к папке для файлов до АОИ
        _after_AOI_folder_path (Optional[Path]): Путь к папке для файлов после АОИ
        _defective_dice_folder_path (Optional[Path]): Путь к папке для хранения изображения кристаллов
        _before_AOI_excel_protocol_file_path (Optional[Path]): Путь к Excel-файлу карты годности до АОИ
        _after_AOI_excel_protocol_file_path (Optional[Path]): Путь к Excel-файлу карты годности после АОИ
        _before_AOI_bin_protocol_file_path (Optional[Path]): Путь к бинарнику карты годности до АОИ
        _after_AOI_bin_protocol_file_path (Optional[Path]): Путь к бинарнику карты годности после АОИ
        _json_protocol_file_path (Optional[Path]): Путь к json-файлу с текстовым видом протокола
        _symbol_colors (Optional[Dict[str, str]]): Словарь цветов для символов
        _initialized (bool): Флаг инициализации
        is_success_save_photos (bool): Флаг успешности сохранения фото
        count_need_focus (int): Количество кристаллов, потребовавших автофокусировку
        count_need_centering (int): Количество кристаллов, потребовавших автоцентровку
        _total_inspection_seconds (float): Общее накопленное время инспекции в секундах
        _total_checked_dice (int): Общее количество проверенных кристаллов
        _session_start_time (Optional[float]): Время начала текущей сессии (None если не активна)
    """
    _instance: Optional['Protocol'] = None
    _lock: Lock = Lock()
    _file_lock: Lock = Lock()  # Блокировка для файловых операций

    def __new__(cls) -> 'Protocol':
        """
        Контролирует создание экземпляра класса.
        Если экземпляр уже существует, возвращает его, иначе создает новый.

        Returns:
            Protocol: Единственный экземпляр класса
        """
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self) -> None:
        """ Инициализация атрибутов экземпляра. """
        # Защита от повторной инициализации
        if hasattr(self, '_initialized') and self._initialized:
            return

        self._main_folder_path = None
        self._before_AOI_folder_path = None
        self._after_AOI_folder_path = None
        self._defective_dice_folder_path = None
        self._before_AOI_excel_protocol_file_path = None
        self._after_AOI_excel_protocol_file_path = None
        self._before_AOI_bin_protocol_file_path = None
        self._after_AOI_bin_protocol_file_path = None
        self._json_protocol_file_path = None
        self._symbol_colors = None

        self._initialized = False
        self.is_success_save_photos = True

        self.count_need_focus: int = 0
        self.count_need_centering: int = 0

        # Упрощенная логика времени: только общее время и время начала сессии
        self._total_inspection_seconds: float = 0.0  # Общее накопленное время
        self._total_checked_dice: int = 0  # Общее количество проверенных кристаллов
        self._session_start_time: Optional[float] = None  # Время начала текущей сессии (None если не активна)

        atexit.register(self._save_on_exit)

    def _save_protocol_data(self) -> bool:
        """
        Единый метод для сохранения данных протокола.

        Returns:
            bool: True если сохранение успешно, иначе False
        """
        try:
            if not WaferMap.has_instance():
                return False

            wafer_map = WaferMap.get_instance()
            if wafer_map is None or not self._initialized:
                return False

            self.update_protocol(wafer_map)
            return True

        except Exception as e:
            logger.error(f"Ошибка при сохранении протокола: {e}")
            return False

    def _save_on_exit(self) -> None:
        """Вызывается при завершении программы через atexit."""
        self._save_protocol_data()

    def shutdown(self) -> None:
        """Метод корректного завершения работы протокола. Сохраняет все данные перед закрытием."""
        self._save_protocol_data()

    def __del__(self):
        """Деструктор класса Protocol."""
        self._save_protocol_data()

    @classmethod
    def create_new_instance(cls) -> 'Protocol':
        """
        Создает новый экземпляр Protocol, заменяя старый.

        Returns:
            Protocol: Новый экземпляр класса
        """
        with cls._lock:
            cls._instance = None
            new_instance = cls()
            return new_instance

    @classmethod
    def get_instance(cls) -> 'Protocol':
        """
        Получение единственного экземпляра Protocol.

        Returns:
            Protocol: Единственный экземпляр класса
        """
        if cls._instance is None:
            cls()
        return cls._instance

    @classmethod
    def has_instance(cls) -> bool:
        """
        Проверяет, создан ли уже экземпляр Protocol.

        Returns:
            bool: True если экземпляр существует, иначе False
        """
        return cls._instance is not None

    @classmethod
    def reset_instance(cls) -> None:
        """
        Сбрасывает Singleton экземпляр.
        Используется в основном для тестирования.
        """
        with cls._lock:
            if cls._instance is not None:
                cls._instance = None

    def _init(self) -> None:
        """Инициализация атрибутов класса из JSON-файла."""
        json_file_path = "project/configuration/protocol.json"

        def get_path(key: str) -> Optional[Path]:
            value = read_from_json(json_file_path, key)
            return Path(value) if value else None

        self._main_folder_path = get_path("main_folder_path")
        self._after_AOI_folder_path = get_path("after_AOI_folder_name")
        self._before_AOI_folder_path = get_path("before_AOI_folder_name")
        self._defective_dice_folder_path = get_path("defective_dice_folder_name")
        self._after_AOI_excel_protocol_file_path = get_path("after_AOI_excel_protocol_file_name")
        self._before_AOI_excel_protocol_file_path = get_path("before_AOI_excel_protocol_file_name")
        self._json_protocol_file_path = get_path("json_protocol_file_name")
        self._symbol_colors = read_from_json(json_file_path, "symbol_colors")

        logger.debug("Информация для протоколов изъята из конфигурационного файла")
        if self._is_inited():
            self._initialized = True

    def _is_inited(self, is_check_bin_paths: bool = False) -> bool:
        """
        Проверка инициализации класса - проверяет, что все поля не None.

        Args:
            is_check_bin_paths: Флаг проверки путей к бинарным файлам

        Returns:
            bool: True если все поля инициализированы, иначе False
        """
        required_fields = {
            "main_folder_path": self._main_folder_path,
            "after_AOI_folder_path": self._after_AOI_folder_path,
            "before_AOI_folder_path": self._before_AOI_folder_path,
            "defective_dice_folder_path": self._defective_dice_folder_path,
            "after_AOI_excel_protocol_file_path": self._after_AOI_excel_protocol_file_path,
            "before_AOI_excel_protocol_file_path": self._before_AOI_excel_protocol_file_path,
            "json_protocol_file_path": self._json_protocol_file_path,
            "symbol_colors": self._symbol_colors
        }
        if is_check_bin_paths:
            required_fields.update({
                "before_AOI_bin_protocol_file_path": self._before_AOI_bin_protocol_file_path,
                "after_AOI_bin_protocol_file_path": self._after_AOI_bin_protocol_file_path,
            })

        for field_name, field_value in required_fields.items():
            if field_value is None:
                logger.debug(f"Поле '{field_name}' = None")

        return all(field is not None for field in required_fields.values())

    # ==================== УПРОЩЕННЫЕ МЕТОДЫ УПРАВЛЕНИЯ ВРЕМЕНЕМ ====================

    def start_timer(self) -> None:
        """
        Начинает отсчет времени инспекции.
        Вызывается при старте инспекции или при продолжении после паузы.
        Если таймер уже запущен, ничего не делает.
        """
        if self._session_start_time is None:
            self._session_start_time = time.time()
            logger.debug(f"Таймер инспекции запущен в {self._session_start_time}")
        else:
            logger.debug("Таймер уже запущен, повторный запуск игнорируется")

    def stop_timer(self) -> None:
        """
        Останавливает отсчет времени и добавляет прошедшее время к общему.
        Вызывается при остановке инспекции или при паузе.
        Если таймер не запущен, ничего не делает.
        """
        if self._session_start_time is not None:
            elapsed = time.time() - self._session_start_time
            self._total_inspection_seconds += elapsed
            self._session_start_time = None
            logger.debug(f"Таймер остановлен. Добавлено {elapsed:.3f} сек. "
                         f"Общее время: {self._total_inspection_seconds:.3f} сек")
        else:
            logger.debug("Таймер не был запущен, остановка игнорируется")

    def _get_average_time_per_die(self) -> Optional[float]:
        """
        Возвращает среднее время на кристалл.

        Returns:
            Optional[float]: Среднее время в секундах или None если нет данных
        """
        if self._total_checked_dice > 0:
            return self._total_inspection_seconds / self._total_checked_dice
        return None

    def _format_inspection_time(self) -> str:
        """
        Форматирует общее время инспекции в читаемый формат.

        Returns:
            str: Время в формате "ЧЧ:ММ:СС" или "ММ:СС" в зависимости от длительности
        """
        total_seconds = self._total_inspection_seconds
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)

        if hours > 0:
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        else:
            return f"{minutes:02d}:{seconds:02d}"

    def _update_checked_dice_count(self, count: int) -> None:
        """
        Обновляет количество проверенных кристаллов.

        Args:
            count: Новое количество проверенных кристаллов
        """
        self._total_checked_dice = count
        logger.debug(f"Обновлено количество проверенных кристаллов: {count}")

    # ==================== МЕТОДЫ СОЗДАНИЯ И ОБНОВЛЕНИЯ ПРОТОКОЛОВ ====================

    def create_protocol(self,
                        protocol_path: str,
                        wafer_map_bin_file_path: str,
                        wafer_map: 'WaferMap') -> Path:
        """
        Создает протокол на основе уже созданного WaferMap.

        Args:
            protocol_path: Путь к текущему протоколу
            wafer_map_bin_file_path: Путь к бинарнику с картой годности
            wafer_map: Объект WaferMap (модель данных пластины с кристаллами)

        Returns:
            Path: Путь к корневой папке протокола

        Raises:
            ProtocolException: При ошибках создания протоколов
        """
        self._reset()
        self._init()

        if not self._initialized:
            error_message = "Отсутствуют названия папок с протоколами. Создание файлов не возможно"
            logger.error(error_message)
            raise ProtocolException(message=error_message)

        current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        wafer_id = wafer_map.wafer_id or ""

        if not wafer_id:
            wafer_id = current_time
            logger.warning(f"Wafer ID не найден. Используется дата: {wafer_id}")
        else:
            wafer_id = f"{wafer_id}_{current_time}"

        # Определение путей к папкам (используем protocol_path как базовый)
        self._main_folder_path = Path(protocol_path) / f"{self._main_folder_path.name}_{wafer_id}"
        self._before_AOI_folder_path = self._main_folder_path / self._before_AOI_folder_path.name
        self._after_AOI_folder_path = self._main_folder_path / self._after_AOI_folder_path.name
        self._defective_dice_folder_path = self._after_AOI_folder_path / self._defective_dice_folder_path.name

        json_name = self._json_protocol_file_path.stem
        self._json_protocol_file_path = (self._after_AOI_folder_path /
                                         f"{json_name}_{wafer_id}{self._json_protocol_file_path.suffix}")

        excel_before_name = self._before_AOI_excel_protocol_file_path.stem
        self._before_AOI_excel_protocol_file_path = (
                self._before_AOI_folder_path
                / f"{excel_before_name}_{wafer_id}{self._before_AOI_excel_protocol_file_path.suffix}")

        excel_after_name = self._after_AOI_excel_protocol_file_path.stem
        self._after_AOI_excel_protocol_file_path = (
                self._after_AOI_folder_path
                / f"{excel_after_name}_{wafer_id}{self._after_AOI_excel_protocol_file_path.suffix}")

        bin_source_path = Path(wafer_map_bin_file_path)
        bin_name = bin_source_path.stem
        bin_suffix = bin_source_path.suffix

        self._before_AOI_bin_protocol_file_path = self._before_AOI_folder_path / f"{bin_name}{bin_suffix}"
        self._after_AOI_bin_protocol_file_path = self._after_AOI_folder_path / f"{bin_name}_udp{bin_suffix}"

        try:
            self._main_folder_path.mkdir(parents=True, exist_ok=True)
            self._before_AOI_folder_path.mkdir(exist_ok=True)
            self._after_AOI_folder_path.mkdir(exist_ok=True)
            self._defective_dice_folder_path.mkdir(exist_ok=True)
            logger.debug(f"Созданы все необходимые папки для протоколов")
        except Exception as e:
            error_message = "Ошибка создания папок с протоколами"
            logger.error(f"{error_message}: {e}")
            raise ProtocolException(message=error_message)

        # Копирование исходного бинарника в протоколы (до АОИ)
        try:
            shutil.copy2(wafer_map_bin_file_path, self._before_AOI_bin_protocol_file_path)
            logger.debug(f"Исходный bin-файл с картой годности скопирован в папку 'До_АОИ': "
                         f"{self._before_AOI_bin_protocol_file_path}")

            shutil.copy2(wafer_map_bin_file_path, self._after_AOI_bin_protocol_file_path)
            logger.debug(f"Исходный bin-файл с картой годности скопирован в папку 'После_АОИ': "
                         f"{self._after_AOI_bin_protocol_file_path}")

        except Exception as e:
            logger.warning(f"Не удалось скопировать файл: {e}")

        # Сброс статистики времени при создании нового протокола
        self._total_inspection_seconds = 0.0
        self._total_checked_dice = 0
        self._session_start_time = None

        # Создание json-файла на основе wafer_map
        self._create_json_protocol(wafer_map)

        # Создание Excel-файлов на основе wafer_map
        self._create_excel_protocol_files(wafer_map)
        logger.info(f"Структура протокольных папок со всеми необходимыми данными успешно создана")

        return self._main_folder_path

    def _create_json_protocol(self,
                              wafer_map: 'WaferMap',
                              json_file_path: Optional[Path] = None) -> bool:
        """
        Создаёт полную структуру json-отчёт со всеми данными.

        Args:
            wafer_map: Объект WaferMap с данными пластины
            json_file_path: Путь к файлу json-файлу

        Returns:
            bool: True в случае успешного создания json-отчёта, иначе False

        Raises:
            ProtocolException: При ошибке создания JSON протокола
        """
        if json_file_path is None:
            json_file_path = self._json_protocol_file_path

        json_file_path = Path(json_file_path)

        total_pass_dice = wafer_map.get_total_pass_dice()
        total_fail_dice = wafer_map.get_total_fail_dice()

        stats = wafer_map.get_stats()
        symbols_dict = {}
        for symbol, count in stats.items():
            if count > 0:
                symbols_dict[symbol] = count

        # Собираем статистику по дефектам
        defects_stats = self._collect_defects_statistics(wafer_map)

        # Формируем список кристаллов после инспекции (все кристаллы на текущий момент)
        dices_info = []
        for row_idx in range(wafer_map.total_rows):
            for col_idx in range(wafer_map.total_cols):
                die = wafer_map.die_matrix[row_idx][col_idx]
                if die:
                    # Формируем информацию о дефектах для кристалла
                    defects_list = []
                    if die.defects_info and len(die.defects_info) > 0:
                        for defect in die.defects_info:
                            defect_data = {
                                "name": defect.get('name', 'Неизвестный дефект'),
                                "color": defect.get('color', [255, 255, 255]),
                                "count": defect.get('count', 0)
                            }
                            defects_list.append(defect_data)

                    die_info = {
                        "id": die.id,
                        "map_x": die.col + 1,
                        "map_y": die.row + 1,
                        "die_coordinator_values_x": die.map_x,
                        "die_coordinator_values_y": die.map_y,
                        "physical_x": float(die.physical_x) if die.physical_x is not None else None,
                        "physical_y": float(die.physical_y) if die.physical_y is not None else None,
                        "symbol": die.symbol,
                        "symbol_old": die.symbol_old if die.symbol_old != "" else die.symbol,
                        "status": die.status.value if die.status else None,
                        "file_frame_original_path": str(
                            die.file_frame_original_path) if die.file_frame_original_path else "",
                        "file_frame_filtered_path": str(
                            die.file_frame_filtered_path) if die.file_frame_filtered_path else "",
                        "defects_info": defects_list,
                        "total_defects_on_die": sum(d.get('count', 0) for d in defects_list)
                    }
                    dices_info.append(die_info)

        # Вычисляем среднее время (безопасно)
        avg_time = self._get_average_time_per_die()

        # Базовая структура отчета со всеми полями
        json_data = {
            # Основная информация
            "wafer_id": wafer_map.wafer_id,
            "cell_size_x_mm": wafer_map.cell_size_x_mm,
            "cell_size_y_mm": wafer_map.cell_size_y_mm,
            "total_dice": total_pass_dice + total_fail_dice,

            # Упрощенная логика времени - только общее время и среднее
            "total_inspection_seconds": self._total_inspection_seconds,
            "total_checked_dice": self._total_checked_dice,
            "average_time_per_die": avg_time,  # Может быть None
            "inspection_time_formatted": self._format_inspection_time(),

            # Параметры пластины
            "total_rows": wafer_map.total_rows,
            "total_cols": wafer_map.total_cols,
            "first_die_X": wafer_map.first_die_X,
            "first_die_y": wafer_map.first_die_Y,

            "symbols_need_check": wafer_map.symbols_need_check.copy() if wafer_map.symbols_need_check else [],

            # Пути к папкам и файлам
            "main_folder_path": str(self._main_folder_path),
            "before_AOI_folder_name": self._before_AOI_folder_path.name,
            "after_AOI_folder_name": self._after_AOI_folder_path.name,
            "defective_dice_folder_name": self._defective_dice_folder_path.name,
            "before_AOI_excel_protocol_file_name": self._before_AOI_excel_protocol_file_path.name,
            "after_AOI_excel_protocol_file_name": self._after_AOI_excel_protocol_file_path.name,
            "before_AOI_bin_protocol_file_name": self._before_AOI_bin_protocol_file_path.name,
            "after_AOI_bin_protocol_file_name": self._after_AOI_bin_protocol_file_path.name,
            "json_protocol_file_name": self._json_protocol_file_path.name,

            # Статистика до инспекции
            "start_stats": {
                **symbols_dict,
                "BAD": wafer_map.get_count_dice_of_status(DieStatus.BAD),
                "GOOD": wafer_map.get_count_dice_of_status(DieStatus.GOOD),
                "NEED_CHECK": wafer_map.get_count_dice_of_status(DieStatus.NEED_CHECK),
                "SKIP": wafer_map.get_count_dice_of_status(DieStatus.SKIP),
                "DUMMY": wafer_map.get_count_dice_of_status(DieStatus.DUMMY),
                "total_fail_dice": total_fail_dice,
                "total_pass_dice": total_pass_dice
            },

            # Статистика после инспекции (такая же как start_stats на момент создания)
            "final_stats": {
                **symbols_dict,
                "BAD": wafer_map.get_count_dice_of_status(DieStatus.BAD),
                "GOOD": wafer_map.get_count_dice_of_status(DieStatus.GOOD),
                "NEED_CHECK": wafer_map.get_count_dice_of_status(DieStatus.NEED_CHECK),
                "SKIP": wafer_map.get_count_dice_of_status(DieStatus.SKIP),
                "DUMMY": wafer_map.get_count_dice_of_status(DieStatus.DUMMY),
                "total_fail_dice": total_fail_dice,
                "total_pass_dice": total_pass_dice
            },

            # Статистика по дефектам
            "defects_statistics": defects_stats,
            "total_defects": sum(defects_stats.values()),

            # Список кристаллов после инспекции
            "dices_info": dices_info,

            "die_prev_ref_id": wafer_map.die_prev_ref.id if wafer_map.die_prev_ref else None,
            "count_need_focus": self.count_need_focus,
            "count_need_centering": self.count_need_centering,
        }

        # Используем блокировку для файловых операций
        with self._file_lock:
            try:
                # Создаем родительскую директорию, если её нет
                json_file_path.parent.mkdir(parents=True, exist_ok=True)
                with open(json_file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=2)

                logger.debug(f"JSON протокол создан: {json_file_path}")
                logger.debug(f"Добавлена информация по {len(dices_info)} кристаллам")
                logger.debug(f"symbols_need_check: {wafer_map.symbols_need_check}")

                # Безопасное логирование среднего времени
                avg_time_str = f"{avg_time:.3f}" if avg_time is not None else "N/A"
                logger.debug(f"Время инспекции: {self._format_inspection_time()} "
                             f"(среднее: {avg_time_str} сек/крист)")

                # Логируем количество кристаллов с дефектами
                dice_with_defects = sum(1 for die in dices_info if die.get("defects_info"))
                logger.debug(f"Кристаллов с дефектами: {dice_with_defects}")

                # Логируем статистику по дефектам
                if defects_stats:
                    logger.debug(f"Статистика по дефектам: {defects_stats}")

                return True

            except Exception as e:
                error_message = "Ошибка создания JSON протокола"
                logger.error(f"{error_message}: {e}")
                raise ProtocolException(message=error_message)

    def _update_json_protocol(self,
                              wafer_map: 'WaferMap',
                              json_file_path: Optional[Path] = None) -> bool:
        """
        Обновляет json-отчёт, обновляя существующие поля (final_stats, dices_info, время).

        Args:
            wafer_map: Объект WaferMap с данными пластины
            json_file_path: Путь к файлу json-файлу

        Returns:
            bool: True в случае успешного обновления json-отчёта, иначе False

        Raises:
            ProtocolException: При ошибке обновления JSON протокола
        """
        if json_file_path is None:
            json_file_path = self._json_protocol_file_path

        json_file_path = Path(json_file_path)

        # Используем блокировку для файловых операций
        with self._file_lock:
            try:
                if not json_file_path.exists():
                    error_message = f"JSON файл не найден: {json_file_path}"
                    logger.error(error_message)
                    raise ProtocolException(message=error_message)

                # Загружаем существующий JSON
                with open(json_file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)

                # Обновляем финальную статистику (перезаписываем актуальными данными)
                total_pass_dice = wafer_map.get_total_pass_dice()
                total_fail_dice = wafer_map.get_total_fail_dice()

                stats = wafer_map.get_stats()
                symbols_dict = {}
                for symbol, count in stats.items():
                    if count > 0:
                        symbols_dict[symbol] = count

                json_data["final_stats"] = {
                    **symbols_dict,
                    "BAD": wafer_map.get_count_dice_of_status(DieStatus.BAD),
                    "GOOD": wafer_map.get_count_dice_of_status(DieStatus.GOOD),
                    "NEED_CHECK": wafer_map.get_count_dice_of_status(DieStatus.NEED_CHECK),
                    "SKIP": wafer_map.get_count_dice_of_status(DieStatus.SKIP),
                    "DUMMY": wafer_map.get_count_dice_of_status(DieStatus.DUMMY),
                    "total_fail_dice": total_fail_dice,
                    "total_pass_dice": total_pass_dice
                }

                # Собираем статистику по дефектам
                defects_stats = self._collect_defects_statistics(wafer_map)
                json_data["defects_statistics"] = defects_stats
                json_data["total_defects"] = sum(defects_stats.values())

                # Формируем список кристаллов после инспекции (все кристаллы на текущий момент)
                dices_info = []
                for row_idx in range(wafer_map.total_rows):
                    for col_idx in range(wafer_map.total_cols):
                        die = wafer_map.die_matrix[row_idx][col_idx]
                        if die:
                            # Формируем информацию о дефектах для кристалла
                            defects_list = []
                            if die.defects_info and len(die.defects_info) > 0:
                                for defect in die.defects_info:
                                    defect_data = {
                                        "name": defect.get('name', 'Неизвестный дефект'),
                                        "color": defect.get('color', [255, 255, 255]),
                                        "count": defect.get('count', 0)
                                    }
                                    defects_list.append(defect_data)

                            die_info = {
                                "id": die.id,
                                "map_x": die.col + 1,
                                "map_y": die.row + 1,
                                "die_coordinator_values_x": die.map_x,
                                "die_coordinator_values_y": die.map_y,
                                "physical_x": float(die.physical_x) if die.physical_x is not None else None,
                                "physical_y": float(die.physical_y) if die.physical_y is not None else None,
                                "symbol": die.symbol,
                                "symbol_old": die.symbol_old if die.symbol_old != "" else die.symbol,
                                "status": die.status.value if die.status else None,
                                "file_frame_original_path": str(
                                    die.file_frame_original_path) if die.file_frame_original_path else "",
                                "file_frame_filtered_path": str(
                                    die.file_frame_filtered_path) if die.file_frame_filtered_path else "",
                                "defects_info": defects_list,
                                "total_defects_on_die": sum(d.get('count', 0) for d in defects_list)
                            }
                            dices_info.append(die_info)

                json_data["dices_info"] = dices_info

                json_data["die_prev_ref_id"] = wafer_map.die_prev_ref.id if wafer_map.die_prev_ref else None

                json_data["count_need_focus"] = self.count_need_focus
                json_data["count_need_centering"] = self.count_need_centering

                # Обновляем статистику времени
                avg_time = self._get_average_time_per_die()

                json_data["total_inspection_seconds"] = self._total_inspection_seconds
                json_data["total_checked_dice"] = self._total_checked_dice
                json_data["average_time_per_die"] = avg_time
                json_data["inspection_time_formatted"] = self._format_inspection_time()

                # Сохраняем обновленный JSON
                with open(json_file_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, ensure_ascii=False, indent=2)

                logger.debug(f"JSON протокол обновлён: {json_file_path}")
                logger.debug(f"Обновлена информация по {len(dices_info)} кристаллам")

                # Безопасное логирование среднего времени
                avg_time_str = f"{avg_time:.3f}" if avg_time is not None else "N/A"
                logger.debug(f"Время инспекции: {self._format_inspection_time()} "
                             f"(среднее: {avg_time_str} сек/крист)")

                # Логируем количество кристаллов с дефектами
                dice_with_defects = sum(1 for die in dices_info if die.get("defects_info"))
                logger.debug(f"Кристаллов с дефектами: {dice_with_defects}")

                # Логируем статистику по дефектам
                if defects_stats:
                    logger.debug(f"Статистика по дефектам: {defects_stats}")

                return True

            except json.JSONDecodeError as e:
                error_message = "Ошибка парсинга JSON протокола"
                logger.error(f"{error_message}: {e}")
                raise ProtocolException(message=error_message)
            except Exception as e:
                error_message = "Ошибка обновления JSON протокола"
                logger.error(f"{error_message}: {e}")
                raise ProtocolException(message=error_message)

    def _collect_defects_statistics(self, wafer_map: 'WaferMap') -> Dict[str, int]:
        """
        Собирает статистику по всем типам дефектов из всех кристаллов.

        Args:
            wafer_map: Объект WaferMap с данными пластины

        Returns:
            Dict[str, int]: Словарь {название_дефекта: общее_количество}
        """
        defects_stats = {}

        for row_idx in range(wafer_map.total_rows):
            for col_idx in range(wafer_map.total_cols):
                die = wafer_map.die_matrix[row_idx][col_idx]
                if die and die.defects_info:
                    for defect in die.defects_info:
                        defect_name = defect.get('name', 'Неизвестный дефект')
                        count = defect.get('count', 0)
                        defects_stats[defect_name] = defects_stats.get(defect_name, 0) + count

        return defects_stats

    def _create_excel_protocol_files(self,
                                     wafer_map: 'WaferMap',
                                     before_AOI_path: Optional[Path] = None,
                                     after_AOI_path: Optional[Path] = None) -> bool:
        """
        Создает Excel файлы с сеткой кристаллов для протоколов на основе WaferMap.

        Args:
            wafer_map: Объект WaferMap с данными пластины
            before_AOI_path: Путь к папке "До_АОИ"
            after_AOI_path: Путь к папке "После_АОИ"

        Returns:
            bool: True в случае успешного создания excel-файлов с отчётами, иначе False

        Raises:
            ProtocolException: При ошибке создания Excel файлов
        """
        if before_AOI_path is None:
            before_AOI_path = self._before_AOI_excel_protocol_file_path

        if after_AOI_path is None:
            after_AOI_path = self._after_AOI_excel_protocol_file_path

        try:
            self._create_excel_protocol(wafer_map, before_AOI_path)
            logger.debug(f"Excel файл карты годности до_АОИ создан: {before_AOI_path}")
        except Exception as e:
            error_message = "Ошибка создания Excel файла карты годности"
            logger.error(f"{error_message}: {e}")
            raise ProtocolException(message=error_message)

        try:
            shutil.copy2(before_AOI_path, after_AOI_path)
            logger.debug(f"Excel файл карты годности до_АОИ скопирован в папку После_АОИ: {after_AOI_path}")
            return True
        except Exception as e:
            error_message = "Ошибка создания Excel файла карты годности"
            logger.error(f"{error_message}: {e}")
            raise ProtocolException(message=error_message)

    def _create_excel_protocol(self,
                               wafer_map: 'WaferMap',
                               excel_file_path: Optional[Path] = None) -> bool:
        """
        Создает Excel файл с сеткой кристаллов на основе модели WaferMap.

        Args:
            wafer_map: Объект WaferMap с данными пластины
            excel_file_path: Путь к Excel-файлу

        Returns:
            bool: True в случае успешного создания excel-файла с отчётом, иначе False

        Raises:
            ProtocolException: При ошибке сохранения Excel файла
        """
        if excel_file_path is None:
            excel_file_path = self._before_AOI_excel_protocol_file_path

        HEADER_ROW_OFFSET = 1
        HEADER_COL_OFFSET = 1
        DATA_START_ROW = HEADER_ROW_OFFSET + 1
        DATA_START_COL = HEADER_COL_OFFSET + 1

        col_size = wafer_map.total_cols
        row_size = wafer_map.total_rows

        wb = Workbook()
        ws = wb.active
        ws.title = "Карта годности кристаллов"

        # Стили
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        center_alignment = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(
            start_color="E0E0E0",
            end_color="E0E0E0",
            fill_type="solid"
        )

        header_font = Font(size=10, bold=True)
        cell_font = Font(size=10)

        # Цвет по умолчанию для неопределенных символов
        _symbol_colors = self._symbol_colors.copy() if self._symbol_colors else {}
        default_color = "FFFFFF"

        # Создаем заголовки столбцов (1-я строка, начиная со 2-й колонки)
        for col_idx in range(col_size):
            cell = ws.cell(row=HEADER_ROW_OFFSET, column=col_idx + DATA_START_COL)
            cell.value = col_idx + 1
            cell.alignment = center_alignment
            cell.border = thin_border
            cell.fill = header_fill
            cell.font = header_font

        # Заполняем данные и заголовки строк
        for row_idx in range(row_size):
            # Заголовок строки (1-я колонка)
            header_cell = ws.cell(row=row_idx + DATA_START_ROW, column=HEADER_COL_OFFSET)
            header_cell.value = row_idx + 1
            header_cell.alignment = center_alignment
            header_cell.border = thin_border
            header_cell.fill = header_fill
            header_cell.font = header_font

            for col_idx in range(col_size):
                die = wafer_map.die_matrix[row_idx][col_idx]
                symbol = die.symbol if die and die.symbol else "D"

                cell = ws.cell(row=row_idx + DATA_START_ROW, column=col_idx + DATA_START_COL)
                cell.value = symbol
                cell.alignment = center_alignment
                cell.border = thin_border
                cell.font = cell_font

                # Применяем цвет
                first_char = symbol[0] if symbol else "D"
                fill_color = _symbol_colors.get(first_char, default_color)
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )

        # Угловая ячейка (A1)
        corner_cell = ws.cell(row=HEADER_ROW_OFFSET, column=HEADER_COL_OFFSET)
        corner_cell.border = thin_border
        corner_cell.fill = header_fill

        # Настройка ширины колонок
        column_width = 6
        for col_idx in range(1, col_size + DATA_START_COL):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_width

        # Настройка высоты строк
        row_height = 20
        for row_idx in range(1, row_size + DATA_START_ROW):
            ws.row_dimensions[row_idx].height = row_height

        try:
            wb.save(excel_file_path)
            logger.debug(f"Excel файл карты годности перед АОИ сохранен: {excel_file_path}")
            return True
        except Exception as e:
            error_message = "Ошибка сохранения Excel файла карты годности перед АОИ"
            logger.error(f"{error_message}: {e}")
            raise ProtocolException(message=error_message)

    def _update_excel_protocol(self,
                               wafer_map: 'WaferMap',
                               excel_file_path: Optional[Path] = None) -> bool:
        """
        Обновляет Excel-файл протокола AOI на основе данных пластины.
        Устанавливает символы для всех кристаллов в соответствии с их статусом.

        Args:
            wafer_map: Объект WaferMap с данными о пластине
            excel_file_path: Путь к файлу Excel для обновления

        Returns:
            bool: True если обновление прошло успешно, иначе False

        Raises:
            ProtocolException: При ошибке обновления Excel файла
        """
        if not self._initialized:
            return False

        if not wafer_map or not wafer_map.die_matrix:
            logger.error("WaferMap не содержит данных")
            return False

        if excel_file_path is None:
            excel_file_path = self._after_AOI_excel_protocol_file_path

        try:
            if not excel_file_path.exists():
                error_message = f"Файл {excel_file_path} не найден"
                logger.error(error_message)
                return False

            # Проверяем наличие _symbol_colors
            if not self._symbol_colors:
                logger.error("_symbol_colors не инициализированы")
                return False

            HEADER_ROW_OFFSET = 1
            HEADER_COL_OFFSET = 1
            DATA_START_ROW = HEADER_ROW_OFFSET + 1
            DATA_START_COL = HEADER_COL_OFFSET + 1

            wb = load_workbook(excel_file_path)
            ws = wb.active

            updated_count = 0

            status_to_symbol = {
                DieStatus.BAD: "FV",
                DieStatus.GOOD: "PV",
            }

            for row_idx in range(wafer_map.total_rows):
                for col_idx in range(wafer_map.total_cols):
                    die = wafer_map.die_matrix[row_idx][col_idx]

                    if die is None or die.status not in status_to_symbol:
                        continue

                    symbol = status_to_symbol[die.status]

                    # Проверяем, есть ли цвет для этого символа в _symbol_colors
                    if symbol not in self._symbol_colors:
                        logger.warning(f"Символ '{symbol}' не найден в _symbol_colors")
                        continue

                    fill_color = self._symbol_colors[symbol]

                    cell = ws.cell(
                        row=row_idx + DATA_START_ROW,
                        column=col_idx + DATA_START_COL
                    )

                    old_value = cell.value
                    if old_value == symbol:
                        continue

                    cell.value = symbol
                    cell.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid"
                    )

                    updated_count += 1

            if updated_count > 0:
                wb.save(excel_file_path)
                logger.info(f"Excel файл успешно обновлен: {excel_file_path}")
                logger.info(f"Обновлено ячеек: {updated_count}")

            return True

        except PermissionError:
            error_message = f"Нет доступа к файлу {excel_file_path}. Возможно файл открыт в другой программе."
            logger.error(error_message)
            raise ProtocolException(message=error_message)
        except Exception as e:
            error_message = f"Ошибка при обновлении Excel файла: {e}"
            logger.error(error_message)
            raise ProtocolException(message=error_message)

    def _update_udp_bin_file(self,
                             wafer_map: 'WaferMap',
                             bin_file_path: Optional[Path] = None) -> bool:
        """
        Сохраняет модифицированные данные в бинарный файл.

        Args:
            wafer_map: Объект WaferMap с данными пластины
            bin_file_path: Путь к выходному bin-файлу (udp)

        Returns:
            bool: True если сохранение прошло успешно, иначе False

        Raises:
            ProtocolException: При ошибке обновления bin-файла
        """
        if bin_file_path is None:
            bin_file_path = self._after_AOI_bin_protocol_file_path

        patcher = WaferMapBinPatcher(file_path=bin_file_path)
        count_modified_dice = patcher.patch(wafer_map, target_symbol="FV")
        data_to_save = patcher.new_data

        if data_to_save is None:
            logger.debug(f"Нет данных для сохранения в файл {bin_file_path}")
            return True

        if count_modified_dice == 0:
            logger.debug(f"Нет новых данных для обновления {bin_file_path}")
            return True

        try:
            with open(bin_file_path, 'wb') as f:
                f.write(data_to_save)

            logger.info(f"Бинарный файл успешно создан: {bin_file_path}")
            logger.debug(f"Размер файла: {len(data_to_save)} байт")

            return True

        except PermissionError:
            error_message = f"Нет прав на запись в файл {bin_file_path}"
            logger.error(error_message)
            raise ProtocolException(message=error_message)
        except IOError:
            error_message = f"Ошибка ввода-вывода при сохранении файла {bin_file_path}"
            logger.error(error_message)
            raise ProtocolException(message=error_message)
        except Exception:
            error_message = f"Неизвестная ошибка при сохранении файла {bin_file_path}"
            logger.error(error_message)
            raise ProtocolException(message=error_message)

    def update_protocol(self, wafer_map: 'WaferMap' = None) -> None:
        """
        Обновляет файлы протоколов:
        - json-отчет
        - excel-файл с картой годности
        - выходной bin-файл (udp)

        Args:
            wafer_map: Объект WaferMap с данными пластины
        """
        if wafer_map is not None:
            self.stop_timer()

            wafer_map.update_stats()

            # Обновляем количество проверенных кристаллов
            self._update_checked_dice_count(
                wafer_map.get_count_dice_of_status([DieStatus.GOOD, DieStatus.BAD])
            )

            self._update_json_protocol(wafer_map=wafer_map)
            self._update_excel_protocol(wafer_map=wafer_map)
            self._update_udp_bin_file(wafer_map=wafer_map)

    # ==================== МЕТОДЫ РАБОТЫ С ФОТОГРАФИЯМИ ====================

    def save_die_photos(self,
                        folder_name: str,
                        frame_original: Optional[np.ndarray],
                        frame_filtered: Optional[np.ndarray] = None
                        ) -> Tuple[Optional[str], Optional[str]]:
        """
        Сохраняет две фотографии дефектного кристалла в указанную папку.
        Сохраняет в формате JPG с максимальным качеством 100%.

        Args:
            folder_name: Название папки для фотографий
            frame_original: Оригинальное изображение кристалла (numpy array)
            frame_filtered: Изображение кристалла с фильтрами (numpy array)

        Returns:
            Tuple[Optional[str], Optional[str]]: Кортеж из двух путей (оригинал, с фильтрами).
                                                Если сохранение не удалось - None для соответствующего пути.

        Raises:
            ProtocolException: При ошибках инициализации протоколов
        """
        if not self._initialized:
            error_message = "Отсутствуют названия папок с протоколами. Сохранение фотографий не возможно."
            logger.error(error_message)
            self.is_success_save_photos = False
            return None, None

        if frame_original is None or folder_name is None:
            error_message = "Не все данные были переданы для сохранения изображений"
            logger.error(error_message)
            self.is_success_save_photos = False
            return None, None

        def save_image(frame: Optional[np.ndarray], path: Path, description: str) -> bool:
            """Сохраняет изображение и возвращает True при успехе."""
            if frame is None:
                logger.warning(f"Нет данных для {description} фотографии кристалла {folder_name}")
                return False

            success = cv2.imwrite(str(path), frame, [cv2.IMWRITE_JPEG_QUALITY, 100])
            if success and path.exists() and path.stat().st_size > 0:
                return True

            logger.error(f"Не удалось сохранить {description} фотографию кристалла: {path}")
            self.is_success_save_photos = False
            return False

        defective_dice_folder_path = self._defective_dice_folder_path / folder_name
        try:
            defective_dice_folder_path.mkdir(exist_ok=True)

            current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            file_frame_original_path = defective_dice_folder_path / f"frame_original_{current_time}.jpg"
            file_frame_filtered_path = defective_dice_folder_path / f"frame_filtered_{current_time}.jpg"

            # Сохраняем оригинальное изображение
            original_saved = save_image(frame_original, file_frame_original_path, "оригинальную")
            original_path = str(file_frame_original_path) if original_saved else None

            # Сохраняем изображение с фильтрами
            filtered_saved = save_image(frame_filtered, file_frame_filtered_path, "с фильтрами")
            filtered_path = str(file_frame_filtered_path) if filtered_saved else None

            return original_path, filtered_path

        except PermissionError as e:
            error_message = f"Нет прав на запись в папку {defective_dice_folder_path}"
            logger.error(f"{error_message}: {e}")
            self.is_success_save_photos = False
            return None, None
        except Exception as e:
            error_message = f"Неизвестная ошибка при сохранении фотографий: {e}"
            logger.error(error_message)
            self.is_success_save_photos = False
            return None, None

    def check_flag_update_files_success(self) -> Optional[str]:
        """
        Проверяет успешность обновления протоколов.

        Returns:
            Optional[str]: Перечень ошибок создания протоколов или None в случае их отсутствия
        """
        error_message = None
        if not self.is_success_save_photos:
            error_message = ("Ошибка обновления протоколов:\n"
                             + "Не все фотографии бракованных кристаллов были сохранены.\n"
                             + "Будьте внимательны при просмотре протоколов.")

        self.is_success_save_photos = True
        return error_message

    # ==================== МЕТОДЫ ЗАГРУЗКИ И СБРОСА ====================

    def _reset(self) -> None:
        """Сбрасывает все атрибуты класса к значениям по умолчанию."""
        self._main_folder_path = None
        self._after_AOI_folder_path = None
        self._before_AOI_folder_path = None
        self._defective_dice_folder_path = None
        self._after_AOI_excel_protocol_file_path = None
        self._before_AOI_excel_protocol_file_path = None
        self._before_AOI_bin_protocol_file_path = None
        self._after_AOI_bin_protocol_file_path = None
        self._json_protocol_file_path = None
        self._symbol_colors = None

        self._initialized = False

        self.is_success_save_photos = True

        self.count_need_focus = 0
        self.count_need_centering = 0

        # Сброс статистики времени
        self._total_inspection_seconds = 0.0
        self._total_checked_dice = 0
        self._session_start_time = None

    def load_config_from_json(self, json_file_path: str) -> Path:
        """
        Загружает конфигурацию протоколов из JSON файла и устанавливает атрибуты класса.
        Базовый путь извлекается из местоположения JSON файла (папка на 2 уровня выше).

        Args:
            json_file_path: Путь к JSON файлу с конфигурацией

        Returns:
            Path: Путь к корневой папке протокола

        Raises:
            ProtocolException: При других ошибках загрузки конфигурации
        """
        try:
            json_path = Path(json_file_path)

            with open(json_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)

            expected_keys = [
                "main_folder_path",
                "before_AOI_folder_name",
                "after_AOI_folder_name",
                "defective_dice_folder_name",
                "before_AOI_excel_protocol_file_name",
                "after_AOI_excel_protocol_file_name",
                "before_AOI_bin_protocol_file_name",
                "after_AOI_bin_protocol_file_name",
                "json_protocol_file_name"
            ]

            missing_keys = [key for key in expected_keys if key not in config_data]
            if missing_keys:
                error_message = f"В JSON файле отсутствуют обязательные ключи: {', '.join(missing_keys)}"
                logger.error(error_message)
                raise ProtocolException(message=error_message)

            before_AOI_folder_name = config_data["before_AOI_folder_name"]
            after_AOI_folder_name = config_data["after_AOI_folder_name"]
            defective_dice_folder_name = config_data["defective_dice_folder_name"]

            before_AOI_excel_name = config_data["before_AOI_excel_protocol_file_name"]
            after_AOI_excel_name = config_data["after_AOI_excel_protocol_file_name"]
            before_AOI_bin_name = config_data["before_AOI_bin_protocol_file_name"]
            after_AOI_bin_name = config_data["after_AOI_bin_protocol_file_name"]
            json_protocol_name = config_data["json_protocol_file_name"]

            self._main_folder_path = json_path.parent.parent
            if self._main_folder_path:
                self._before_AOI_folder_path = self._main_folder_path / before_AOI_folder_name
                self._after_AOI_folder_path = self._main_folder_path / after_AOI_folder_name
                self._defective_dice_folder_path = self._after_AOI_folder_path / defective_dice_folder_name

                self._before_AOI_excel_protocol_file_path = self._before_AOI_folder_path / before_AOI_excel_name
                self._after_AOI_excel_protocol_file_path = self._after_AOI_folder_path / after_AOI_excel_name

                self._before_AOI_bin_protocol_file_path = self._before_AOI_folder_path / before_AOI_bin_name
                self._after_AOI_bin_protocol_file_path = self._after_AOI_folder_path / after_AOI_bin_name

                self._json_protocol_file_path = self._after_AOI_folder_path / json_protocol_name

            self._symbol_colors = read_from_json("project/configuration/protocol.json", "symbol_colors")

            if "count_need_focus" in config_data:
                self.count_need_focus = config_data["count_need_focus"]
            if "count_need_centering" in config_data:
                self.count_need_centering = config_data["count_need_centering"]

            # Загружаем статистику времени из JSON
            if "total_inspection_seconds" in config_data:
                self._total_inspection_seconds = config_data["total_inspection_seconds"]
            if "total_checked_dice" in config_data:
                self._total_checked_dice = config_data["total_checked_dice"]

            # Сессия не активна при загрузке
            self._session_start_time = None

            if not self._is_inited(is_check_bin_paths=True):
                error_message = "Не все обязательные поля были загружены из конфигурации"
                logger.error(error_message)
                raise ProtocolException(message=error_message)

            self._initialized = True

            self._validate_paths_existence()  # Проверяем существование всех необходимых файлов и папок

            logger.info(f"Конфигурация протоколов успешно загружена из {json_file_path}")

            # Безопасное логирование среднего времени
            avg_time = self._get_average_time_per_die()
            avg_time_str = f"{avg_time:.3f}" if avg_time is not None else "N/A"
            logger.info(f"Статистика инспекции: время={self._total_inspection_seconds:.3f} сек, "
                        f"кристаллов={self._total_checked_dice}, "
                        f"среднее={avg_time_str} сек/крист")
            return self._main_folder_path

        except FileNotFoundError as e:
            logger.error(f"Файл конфигурации не найден: {e}")
            raise ProtocolException(message="Файл конфигурации не найден")
        except json.JSONDecodeError as e:
            error_message = f"Ошибка парсинга JSON файла {json_file_path}: {e}"
            logger.error(error_message)
            raise ProtocolException(message=error_message)
        except ProtocolException:
            raise
        except Exception as e:
            error_message = f"Неизвестная ошибка при загрузке конфигурации из {json_file_path}: {e}"
            logger.error(error_message)
            raise ProtocolException(message=error_message)

    def _validate_paths_existence(self) -> bool:
        """
        Проверяет существование всех путей, загруженных из конфигурации.
        Выполняет жесткую валидацию: проверяет существование, тип (папка/файл) и формат файлов.

        Returns:
            bool: True в случае правильности путей, иначе False

        Raises:
            ProtocolException: Если какая-либо проверка не пройдена
        """
        problems = []

        # Список папок для проверки: (путь, название, обязательна ли)
        folders_to_check = [
            (self._main_folder_path, self._main_folder_path.name if self._main_folder_path else "main_folder", True),
            (self._before_AOI_folder_path,
             self._before_AOI_folder_path.name if self._before_AOI_folder_path else "before_AOI_folder", True),
            (self._after_AOI_folder_path,
             self._after_AOI_folder_path.name if self._after_AOI_folder_path else "after_AOI_folder", True),
            (self._defective_dice_folder_path,
             self._defective_dice_folder_path.name if self._defective_dice_folder_path else "defective_dice_folder",
             False),
        ]

        for folder_path, folder_name, required in folders_to_check:
            if folder_path is None:
                if required:
                    problems.append(f"Путь '{folder_name}' не определен (None)")
                continue

            if not folder_path.exists():
                problems.append(f"Папка '{folder_name}' не существует: {folder_path}")
            elif not folder_path.is_dir():
                problems.append(f"Путь '{folder_name}' не является папкой: {folder_path}")

        # Список файлов для проверки: (путь, название, обязателен ли, ожидаемое расширение)
        files_to_check = [
            (self._before_AOI_excel_protocol_file_path,
             self._before_AOI_excel_protocol_file_path.name if self._before_AOI_excel_protocol_file_path
             else "before_AOI_excel", True, ['.xlsx']),
            (self._after_AOI_excel_protocol_file_path,
             self._after_AOI_excel_protocol_file_path.name if self._after_AOI_excel_protocol_file_path
             else "after_AOI_excel", True, ['.xlsx']),
            (self._before_AOI_bin_protocol_file_path,
             self._before_AOI_bin_protocol_file_path.name if self._before_AOI_bin_protocol_file_path
             else "before_AOI_bin", True, ['.bin', '']),
            (self._after_AOI_bin_protocol_file_path,
             self._after_AOI_bin_protocol_file_path.name if self._after_AOI_bin_protocol_file_path
             else "after_AOI_bin", True, ['.bin', '']),
            (self._json_protocol_file_path,
             self._json_protocol_file_path.name if self._json_protocol_file_path
             else "json_protocol", True, ['.json']),
        ]

        for file_path, file_name, required, expected_extensions in files_to_check:
            if file_path is None:
                if required:
                    problems.append(f"Файл '{file_name}' не определен (None)")
                continue

            # Проверяем существование родительской директории
            parent_dir = file_path.parent
            if not parent_dir.exists():
                problems.append(f"Родительская папка для файла '{file_name}' не существует: {parent_dir}")
                continue

            # Проверяем существование файла (только для обязательных)
            if required and not file_path.exists():
                problems.append(f"Обязательный файл '{file_name}' не существует: {file_path}")
                continue

            # Проверяем расширение файла
            if file_path.suffix.lower() not in expected_extensions:
                problems.append(f"Файл '{file_name}' имеет неверное расширение. "
                                f"Ожидается: {', '.join(expected_extensions)}, "
                                f"Получено: {file_path.suffix}")

        # Проверяем, что defective_dice_folder является подпапкой after_AOI_folder
        if self._defective_dice_folder_path and self._after_AOI_folder_path:
            try:
                # Проверяем, что defective_dice_folder находится внутри after_AOI_folder
                self._defective_dice_folder_path.relative_to(self._after_AOI_folder_path)
            except ValueError:
                problems.append(f"defective_dice_folder не является подпапкой after_AOI_folder: "
                                f"{self._defective_dice_folder_path} не содержит {self._after_AOI_folder_path}")

        # Проверяем, что все пути имеют правильную структуру (не содержат двойных слешей и т.д.)
        all_paths = folders_to_check + files_to_check
        for path_obj, path_name, required, *args in all_paths:
            if path_obj and isinstance(path_obj, Path):
                # Проверяем, что путь абсолютный
                if not path_obj.is_absolute():
                    problems.append(f"Путь '{path_name}' не является абсолютным: {path_obj}")

                # Проверяем, что в пути нет None или пустых компонентов
                for part in path_obj.parts:
                    if not part or part == '.' or part == '..':
                        problems.append(f"Путь '{path_name}' содержит некорректные компоненты: {path_obj}")
                        break

        if problems:
            error_message = "Ошибки валидации путей:\n" + "\n".join(f"  - {p}" for p in problems)
            logger.error(error_message)
            raise ProtocolException(message=error_message)

        logger.debug("Все пути в конфигурации валидны")
        return True
